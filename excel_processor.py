"""
Excel-specific functionality for Spotify ISRC Matcher
Handles reading ISRCs from Excel files and saving results to Excel
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Optional
import logging
from datetime import datetime
from spotify_isrc_matcher import SpotifyISRCMatcher, TrackInfo

logger = logging.getLogger(__name__)

class ExcelISRCProcessor:
    def __init__(self, client_id: str, client_secret: str):
        self.matcher = SpotifyISRCMatcher(client_id, client_secret)
    
    def read_isrcs_from_excel(self, excel_path: str, sheet_name: str = None, 
                             isrc_column: str = 'ISRC') -> List[str]:
        """
        Read ISRCs from an Excel file
        
        Args:
            excel_path: Path to the Excel file
            sheet_name: Name of the sheet to read (if None, reads first sheet)
            isrc_column: Name of the column containing ISRCs
            
        Returns:
            List of ISRCs
        """
        try:
            # Read Excel file
            if sheet_name:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(excel_path)
            
            logger.info(f"Successfully read Excel file: {excel_path}")
            logger.info(f"Sheet contains {len(df)} rows")
            
            # Find ISRC column (case-insensitive)
            isrc_col = None
            for col in df.columns:
                if col.upper() == isrc_column.upper():
                    isrc_col = col
                    break
            
            if isrc_col is None:
                # Try common variations
                common_names = ['ISRC', 'isrc', 'ISRC_CODE', 'isrc_code', 'Code', 'code']
                for name in common_names:
                    if name in df.columns:
                        isrc_col = name
                        break
            
            if isrc_col is None:
                raise ValueError(f"Could not find ISRC column. Available columns: {list(df.columns)}")
            
            # Extract ISRCs and clean them
            isrcs = []
            for idx, value in df[isrc_col].items():
                if pd.notna(value):
                    isrc = str(value).strip()
                    if isrc:
                        isrcs.append(isrc)
            
            logger.info(f"Found {len(isrcs)} valid ISRCs in column '{isrc_col}'")
            return isrcs
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise
    
    def save_results_to_excel(self, results: List[TrackInfo], output_path: str, 
                             include_metadata: bool = True):
        """
        Save results to Excel file with formatting
        
        Args:
            results: List of TrackInfo objects
            output_path: Path for the output Excel file
            include_metadata: Whether to include metadata sheet
        """
        try:
            # Prepare data for main results sheet
            data = []
            for track_info in results:
                row = {
                    'ISRC': track_info.isrc,
                    'Release Year': track_info.release_year,
                    'Track Name': track_info.track_name,
                    'Artist Name': track_info.artist_name,
                    'Album Name': track_info.album_name,
                    'Status': 'Success' if track_info.error is None else 'Failed',
                    'Error': track_info.error if track_info.error else ''
                }
                data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Create Excel writer with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write main results
                df.to_excel(writer, sheet_name='Results', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Results']
                
                # Format the results sheet
                self._format_results_sheet(workbook, worksheet, df)
                
                # Add metadata sheet if requested
                if include_metadata:
                    self._add_metadata_sheet(writer, results)
                
                # Add summary sheet
                self._add_summary_sheet(writer, results)
            
            logger.info(f"Results saved to Excel file: {output_path}")
            
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
            raise
    
    def _format_results_sheet(self, workbook, worksheet, df):
        """Apply formatting to the results sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Apply header formatting
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for col in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in range(1, len(df) + 2):
                cell_value = str(worksheet.cell(row=row, column=col).value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Color-code status column
        status_col = 6  # Status column
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=status_col)
            if cell.value == "Success":
                cell.fill = success_fill
            elif cell.value == "Failed":
                cell.fill = error_fill
    
    def _add_metadata_sheet(self, writer, results: List[TrackInfo]):
        """Add metadata sheet with processing information"""
        metadata = {
            'Processing Information': [
                f"Processing Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"Total ISRCs Processed: {len(results)}",
                f"Successful: {sum(1 for r in results if r.error is None)}",
                f"Failed: {sum(1 for r in results if r.error is not None)}",
                f"Success Rate: {(sum(1 for r in results if r.error is None) / len(results) * 100):.1f}%"
            ]
        }
        
        # Common errors
        errors = [r.error for r in results if r.error]
        if errors:
            error_counts = {}
            for error in errors:
                error_counts[error] = error_counts.get(error, 0) + 1
            
            metadata['Common Errors'] = [f"{error}: {count}" for error, count in error_counts.items()]
        
        # Create metadata DataFrame
        max_length = max(len(v) for v in metadata.values())
        metadata_df = pd.DataFrame({k: v + [''] * (max_length - len(v)) for k, v in metadata.items()})
        
        metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
    
    def _add_summary_sheet(self, writer, results: List[TrackInfo]):
        """Add summary sheet with statistics"""
        # Year distribution
        years = [r.release_year for r in results if r.release_year]
        if years:
            year_counts = {}
            for year in years:
                year_counts[year] = year_counts.get(year, 0) + 1
            
            year_df = pd.DataFrame(list(year_counts.items()), columns=['Year', 'Count'])
            year_df = year_df.sort_values('Year')
            year_df.to_excel(writer, sheet_name='Year Distribution', index=False)
        
        # Artist distribution (top 20)
        artists = [r.artist_name for r in results if r.artist_name]
        if artists:
            artist_counts = {}
            for artist in artists:
                artist_counts[artist] = artist_counts.get(artist, 0) + 1
            
            top_artists = sorted(artist_counts.items(), key=lambda x: x[1], reverse=True)[:20]
            artist_df = pd.DataFrame(top_artists, columns=['Artist', 'Track Count'])
            artist_df.to_excel(writer, sheet_name='Top Artists', index=False)
    
    def process_excel_file(self, input_path: str, output_path: str = None, 
                          sheet_name: str = None, isrc_column: str = 'ISRC',
                          delay: float = 0.1) -> List[TrackInfo]:
        """
        Complete workflow: read Excel, process ISRCs, save results
        
        Args:
            input_path: Path to input Excel file
            output_path: Path for output Excel file (auto-generated if None)
            sheet_name: Sheet name to read from
            isrc_column: Column name containing ISRCs
            delay: Delay between API calls
            
        Returns:
            List of TrackInfo objects
        """
        logger.info(f"Starting Excel processing workflow for: {input_path}")
        
        # Read ISRCs from Excel
        isrcs = self.read_isrcs_from_excel(input_path, sheet_name, isrc_column)
        
        # Process ISRCs
        results = self.matcher.process_isrc_list(isrcs, delay)
        
        # Generate output path if not provided
        if output_path is None:
            input_path_obj = Path(input_path)
            output_path = str(input_path_obj.parent / f"{input_path_obj.stem}_results.xlsx")
        
        # Save results to Excel
        self.save_results_to_excel(results, output_path)
        
        logger.info(f"Excel processing complete. Results saved to: {output_path}")
        
        return results

def create_sample_excel():
    """Create a sample Excel file for testing"""
    sample_data = {
        'ISRC': [
            'USUG11904257',  # Blinding Lights
            'GBUM71029604',  # Someone Like You
            'USUM71703861',  # Shape of You
            'USRC17607839',  # Despacito
            'GBAHS1700133',  # Watermelon Sugar
            'INVALID_ISRC',  # Test invalid
        ],
        'Song Title': [
            'Blinding Lights',
            'Someone Like You',
            'Shape of You',
            'Despacito',
            'Watermelon Sugar',
            'Test Invalid'
        ],
        'Notes': [
            'The Weeknd hit',
            'Adele classic',
            'Ed Sheeran popular',
            'Luis Fonsi ft. Daddy Yankee',
            'Harry Styles',
            'Testing error handling'
        ]
    }
    
    df = pd.DataFrame(sample_data)
    sample_path = 'sample_isrcs.xlsx'
    df.to_excel(sample_path, index=False)
    
    print(f"Created sample Excel file: {sample_path}")
    return sample_path

def main():
    """Example usage of Excel processing"""
    import json
    
    # Load configuration
    with open('config.json', 'r') as f:
        config = json.load(f)
    
    # Create sample Excel file if it doesn't exist
    sample_path = 'sample_isrcs.xlsx'
    if not Path(sample_path).exists():
        create_sample_excel()
    
    # Initialize processor
    processor = ExcelISRCProcessor(config['client_id'], config['client_secret'])
    
    # Process the Excel file
    results = processor.process_excel_file(
        input_path=sample_path,
        output_path='spotify_results.xlsx',
        isrc_column='ISRC'
    )
    
    # Print summary
    successful = sum(1 for r in results if r.error is None)
    failed = len(results) - successful
    
    print(f"\n📊 Processing Summary:")
    print(f"   Input file: {sample_path}")
    print(f"   Output file: spotify_results.xlsx")
    print(f"   Total ISRCs: {len(results)}")
    print(f"   Successful: {successful}")
    print(f"   Failed: {failed}")
    print(f"   Success rate: {(successful / len(results) * 100):.1f}%")

if __name__ == "__main__":
    main()
